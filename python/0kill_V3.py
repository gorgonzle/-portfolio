import pandas as pd
from tkinter import Tk, filedialog, messagebox, Checkbutton, BooleanVar
import tkinter as tk
import os

def validate_file_selection(file_paths):
    required_files = ["ELENUM", "CREQV", "PRSTMAX", "VOL"]
    selected_files = {os.path.splitext(os.path.basename(fp))[0]: fp for fp in file_paths}

    missing_files = [file for file in required_files if file not in selected_files]

    if missing_files:
        messagebox.showerror(
            "Missing Files",
            f"The following required files are missing: {', '.join(missing_files)}. Please select them."
        )
        return None

    return [selected_files[file] for file in required_files]

def preprocess_and_save(file_paths):
    """Preprocess files, check consistency, and return processed data."""
    processed_data = {}
    lengths = []

    for file_path in file_paths:
        try:
            data = pd.read_csv(file_path, header=None)  # Use header=None because the CSV files do not have headers

            # Remove rows where all values in all columns are exactly 0
            data = data.loc[~(data == 0).all(axis=1)]

            # Trim whitespace from all string data
            data = data.applymap(lambda x: x.strip() if isinstance(x, str) else x)

            file_name = os.path.splitext(os.path.basename(file_path))[0]

            # Add file name as the first row, shifting all data down
            file_name_row = pd.DataFrame([[file_name] + [""] * (data.shape[1] - 1)])
            data = pd.concat([file_name_row, data], ignore_index=True)

            processed_data[file_name] = data
            lengths.append(len(data))
        except Exception as e:
            return None, f"Failed to process {file_path}: {e}"

    if len(set(lengths)) > 1:
        return None, "Data lengths are inconsistent among the files."

    return processed_data, None

def save_to_excel(processed_data, sheet_name, directory):
    """Save processed data to an Excel sheet, assigning each data to specific columns based on file names."""
    excel_path = os.path.join(directory, "Data_tot.xlsx")

    if not os.path.exists(excel_path):
        messagebox.showerror("Missing Excel File", f"The file 'Data_tot.xlsx' is missing in the directory: {directory}")
        return False

    try:
        from openpyxl import load_workbook

        # Load existing workbook
        workbook = load_workbook(excel_path)

        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)

        column_map = {"ELENUM": "A", "CREQV": "B", "PRSTMAX": "C", "VOL": "D"}  # Map file names to columns

        for file_key, data in processed_data.items():
            if file_key not in column_map:
                messagebox.showerror("Column Mapping Error", f"No column mapping found for {file_key}.")
                return False

            column_letter = column_map[file_key]

            for row_index, value in enumerate(data.iloc[:, 0], start=1):
                sheet[f"{column_letter}{row_index}"] = value

        workbook.save(excel_path)
        return True

    except Exception as e:
        messagebox.showerror("Excel Save Error", f"Failed to save to Excel: {e}")
        return False

def select_and_process_files():
    """Handle file selection and processing."""
    file_paths = filedialog.askopenfilenames(
        title="Select CSV Files",
        filetypes=[("CSV files", "*.csv")]
    )

    validated_files = validate_file_selection(file_paths)
    if not validated_files:
        return

    # Get directory of the first selected file
    directory = os.path.dirname(validated_files[0])

    # Check that a sheet is selected
    if not any(sheet_vars.values()):
        messagebox.showerror("Sheet Not Selected", "Please select a sheet before proceeding.")
        return

    selected_sheet = next(sheet for sheet, var in sheet_vars.items() if var.get())

    processed_data, error = preprocess_and_save(validated_files)
    if error:
        messagebox.showerror("Processing Error", error)
        return

    if save_to_excel(processed_data, selected_sheet, directory):
        for file in file_paths:
            os.remove(file)
        # Remove all 0killed files after successful processing
        for file_key in processed_data.keys():
            output_path = os.path.join(directory, f"{file_key}_0killed.csv")
            if os.path.exists(output_path):
                os.remove(output_path)
        messagebox.showinfo("Processing Complete", "Files processed and saved successfully. Original and 0killed files deleted.")

def create_gui():
    """Create the GUI application."""
    root = Tk()
    root.title("Batch CSV Preprocessor with Excel Integration")
    root.geometry("600x400")

    # Title
    tk.Label(root, text="Batch CSV Preprocessor", font=("Arial", 16)).pack(pady=10)

    # Instructions
    tk.Label(
        root, 
        text=(
            "Select the required CSV files to process and save the results to the Excel file 'Data_tot.xlsx'.\n"
            "Files will be processed and added to the selected sheet. Ensure Excel file exists in the same directory."
        ),
        font=("Arial", 12)
    ).pack(pady=5)

    # Sheet selection
    tk.Label(root, text="Select Target Sheet:", font=("Arial", 12)).pack(pady=5)
    
    global sheet_vars
    sheet_vars = {sheet: BooleanVar() for sheet in ["Seal1","Seal2","Seal3", "FFL", "AFL", "EL", "BL"]}

    def toggle_check(var):
        for sheet, sheet_var in sheet_vars.items():
            if sheet_var != var:
                sheet_var.set(False)

    for sheet, var in sheet_vars.items():
        Checkbutton(root, text=sheet, variable=var, command=lambda v=var: toggle_check(v)).pack(anchor="w")

    # File selection button
    tk.Button(root, text="Select and Process Files", command=select_and_process_files, font=("Arial", 12)).pack(pady=20)

    root.mainloop()

if __name__ == "__main__":
    create_gui()
